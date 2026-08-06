"""Builds the part-by-part incentive workbook from parts-incentive-data.mjs.

Three sheets:
  Parts        every active part, the rule that applies, and what it pays
  Fill costs   the Oasis parts that have no cost yet, ready to be filled in
  Rules        the rules in plain English

Usage, from backend/:
  node --env-file=.env scripts/parts-incentive-data.mjs > parts.json
  python scripts/make-parts-excel.py parts.json "d:/.../Parts-Incentive.xlsx"
"""
import json
import sys
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = sys.argv[1] if len(sys.argv) > 1 else "parts.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Parts-Incentive.xlsx"

data = json.load(open(SRC, encoding="utf-8"))
rows, rules = data["rows"], data["rules"]
ist = datetime.now(timezone(timedelta(hours=5, minutes=30)))

INK = "FF141B17"
HEAD_BG = "FF1F3A2E"
BAND = "FFF2F5F1"
WARN_BG = "FFFBE9E6"
WARN_FG = "FFA32B1C"
MUTED = "FF6B7A70"

thin = Side(style="thin", color="FFD8DFD9")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
RUPEE = '"\u20b9"#,##0.00'


def header(ws, labels, widths):
    ws.append(labels)
    for i, (lab, w) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, size=9, color="FFFFFFFF", name="Segoe UI")
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="right" if i > 2 else "left")
        c.border = box
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def style_row(ws, r, ncols, money_cols, banded, warn=False):
    for i in range(1, ncols + 1):
        c = ws.cell(row=r, column=i)
        c.border = box
        c.font = Font(size=10, name="Segoe UI",
                      color=WARN_FG if (warn and i in money_cols) else INK,
                      bold=bool(warn and i in money_cols))
        if i in money_cols:
            c.number_format = RUPEE
            c.alignment = Alignment(horizontal="right")
        elif i > 2:
            c.alignment = Alignment(horizontal="right")
        if warn:
            c.fill = PatternFill("solid", fgColor=WARN_BG)
        elif banded:
            c.fill = PatternFill("solid", fgColor=BAND)


wb = Workbook()

# ---------------------------------------------------------------- Parts sheet
ws = wb.active
ws.title = "Parts"
# Cash and online pay the same now, so a third payout column would just repeat
# the second. What still differs is the daily rate — and only for branded parts.
cols = ["Part", "Brand", "MRP", "Given to tech at", "What applies",
        "Pays normally", f"Pays after a ₹{rules['DAILY_TARGET']:,} day"]
header(ws, cols, [38, 12, 12, 17, 24, 15, 20])

order = {"kent": 0, "aquaguard": 1, "oasis": 2, "other": 3}
rows_sorted = sorted(rows, key=lambda r: (order.get(r["brand"], 9), r["name"].lower()))

for n, r in enumerate(rows_sorted):
    ws.append([
        r["name"], r["brand_label"], r["price"],
        "not set" if r["cost_missing"] else r["cost"],
        r["rule"], r["pays_6_cash"], r["pays_10_cash"],
    ])
    style_row(ws, ws.max_row, len(cols), {3, 4, 6, 7}, banded=(n % 2 == 1),
              warn=r["cost_missing"])
    if r["cost_missing"]:
        c = ws.cell(row=ws.max_row, column=4)
        c.number_format = "General"
        c.alignment = Alignment(horizontal="right")

ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

# ---------------------------------------------------------- Fill-costs sheet
missing = [r for r in rows_sorted if r["cost_missing"]]
ws2 = wb.create_sheet("Fill costs")
ws2["A1"] = ("For each part, enter the price we give it to the technician at. He may sell it "
             "anywhere between that and the MRP; whatever he makes above it is his margin.")
ws2["A1"].font = Font(size=10, italic=True, color=MUTED, name="Segoe UI")
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 26
ws2.append([])
ws2.append(["Part", "Brand", "MRP (max he can charge)", "We give it to him at"])
for i, w in enumerate([40, 12, 22, 21], start=1):
    c = ws2.cell(row=3, column=i)
    c.font = Font(bold=True, size=9, color="FFFFFFFF", name="Segoe UI")
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.border = box
    c.alignment = Alignment(horizontal="right" if i > 2 else "left")
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A4"

for n, r in enumerate(missing):
    ws2.append([r["name"], r["brand_label"], r["price"], None])
    rr = ws2.max_row
    for i in range(1, 5):
        c = ws2.cell(row=rr, column=i)
        c.border = box
        c.font = Font(size=10, name="Segoe UI")
        if i in (3, 4):
            c.number_format = RUPEE
            c.alignment = Alignment(horizontal="right")
        if i == 4:
            c.fill = PatternFill("solid", fgColor="FFFFF6D9")  # the cell to type in
        elif n % 2 == 1:
            c.fill = PatternFill("solid", fgColor=BAND)

# ---------------------------------------------------------------- Rules sheet
ws3 = wb.create_sheet("Rules")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 86
ws3["A1"] = "How incentive is worked out"
ws3["A1"].font = Font(bold=True, size=13, name="Segoe UI", color=INK)
ws3.merge_cells("A1:B1")

lines = [
    ("Kent", f"{rules['BRAND_RATE']*100:g}% of the selling price. Becomes "
             f"{rules['BRAND_RATE_BONUS']*100:g}% for the whole day once that technician bills "
             f"\u20b9{rules['DAILY_TARGET']:,} in one day."),
    ("Aquaguard", "Exactly the same as Kent."),
    ("Oasis", "The margin: what he billed minus the price we gave him the part at. He may "
              "bill anything from that price up to MRP. "
              f"{rules['GST_RATE']*100:g}% GST comes off the margin, cash or online."),
    ("No brand set", "Nothing is paid."),
    ("Service charge", f"Nothing is paid \u2014 but it still counts toward the "
                       f"\u20b9{rules['DAILY_TARGET']:,} daily target."),
    ("Cash or online", "Makes no difference to what the technician earns. GST comes off "
                       "the Oasis margin either way."),
    ("Quantity", "Per piece. Two of a part pays twice."),
    ("Which jobs count", "Only jobs marked CLOSED. Days run on Indian time."),
    ("", ""),
    ("Right now", f"{len(missing)} Oasis parts have no price recorded against them. Two things "
                  "follow: the technician keeps almost the whole sale instead of a margin, "
                  "and there is no floor — he could bill a ₹3,250 part at ₹1 and the app "
                  "would allow it. Filling in the 'Fill costs' sheet fixes both."),
]
r = 3
for k, v in lines:
    ws3.cell(row=r, column=1, value=k).font = Font(bold=True, size=10, name="Segoe UI", color=INK)
    c = ws3.cell(row=r, column=2, value=v)
    c.font = Font(size=10, name="Segoe UI", color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.row_dimensions[r].height = 30 if len(v) > 95 else 16
    r += 1

ws3.cell(row=r + 1, column=1, value="Generated").font = Font(size=9, color=MUTED, name="Segoe UI")
ws3.cell(row=r + 1, column=2,
         value=ist.strftime("%d %b %Y, %I:%M %p IST") + f"  \u00b7  {len(rows)} active parts, live from the portal"
         ).font = Font(size=9, color=MUTED, name="Segoe UI")

wb.save(OUT)
print(f"written {OUT}")
print(f"  Parts       {len(rows_sorted)} rows")
print(f"  Fill costs  {len(missing)} Oasis parts awaiting a cost")
